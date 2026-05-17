import re
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from app.parsing.vlp_types import ParsedVlpRow, VlpParseResult

# Spalten-Mapping (normalisierte Header → Feld)
COLUMN_ALIASES: dict[str, list[str]] = {
    "kabel_name": [
        "kabelnummer",
        "kabel",
        "kabelname",
        "name",
        "kabel-nr",
        "kabel nr",
        "leitungsbezeichnung",
    ],
    "laenge_ist": [
        "laenge ist",
        "länge ist",
        "laenge_ist",
        "ist-laenge",
        "ist-länge",
        "istlaenge",
        "länge ist [m]",
        "laenge [m]",
    ],
    "trommelnummer": ["trommelnummer", "trommel", "trommel-nr", "trommelnr"],
    "verlegeart": ["verlegeart", "verlege-art", "art verlegung"],
    "vlp_nummer": ["vlp-nummer", "vlp nummer", "vlp-nr", "vlp", "protokollnummer"],
}


class VlpExcelParser:
    def parse(self, path: str | Path) -> VlpParseResult:
        wb = load_workbook(path, read_only=True, data_only=True)
        result = VlpParseResult()
        ws = wb.active
        if ws is None:
            result.warnings.append("Excel ohne aktives Blatt")
            return result

        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            result.warnings.append("Leere Excel-Datei")
            return result

        col_map = self._map_columns(header_row)
        if "kabel_name" not in col_map:
            result.warnings.append("Spalte für Kabelname nicht gefunden")
            return result

        for idx, row in enumerate(rows_iter, start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            parsed = self._parse_row(row, col_map, idx)
            if parsed.kabel_name:
                result.rows.append(parsed)

        wb.close()
        return result

    def _map_columns(self, header_row: tuple) -> dict[str, int]:
        mapping: dict[str, int] = {}
        for i, cell in enumerate(header_row):
            if cell is None:
                continue
            norm = self._normalize_header(str(cell))
            for field, aliases in COLUMN_ALIASES.items():
                if norm in aliases or any(a in norm for a in aliases):
                    mapping[field] = i
        return mapping

    @staticmethod
    def _normalize_header(text: str) -> str:
        t = text.lower().strip()
        t = t.replace("ä", "ae").replace("ö", "oe").replace("ü", "ue").replace("ß", "ss")
        t = re.sub(r"[^a-z0-9]+", " ", t)
        return re.sub(r"\s+", " ", t).strip()

    def _parse_row(
        self, row: tuple, col_map: dict[str, int], row_index: int
    ) -> ParsedVlpRow:
        def cell(field: str):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        name_raw = cell("kabel_name")
        name = self._normalize_kabel_name(str(name_raw)) if name_raw else None
        return ParsedVlpRow(
            kabel_name=name,
            laenge_ist=self._parse_decimal(cell("laenge_ist")),
            trommelnummer=self._str(cell("trommelnummer")),
            verlegeart=self._str(cell("verlegeart")),
            vlp_nummer=self._str(cell("vlp_nummer")),
            rohdaten={"source": "excel", "row": row_index},
            row_index=row_index,
        )

    @staticmethod
    def _normalize_kabel_name(text: str) -> str | None:
        t = text.strip().upper()
        match = re.search(r"S\d[\w.]*", t, re.IGNORECASE)
        return match.group(0).upper() if match else (t if t else None)

    @staticmethod
    def _parse_decimal(value) -> Decimal | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        s = str(value).strip().replace(",", ".").replace("m", "").strip()
        try:
            return Decimal(s)
        except InvalidOperation:
            return None

    @staticmethod
    def _str(value) -> str | None:
        if value is None:
            return None
        s = str(value).strip()
        return s or None
